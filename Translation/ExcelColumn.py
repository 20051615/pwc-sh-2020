from Baidu_API import translate_list, prompt_for_target_lang
import openpyxl

def load_xl():
    wb = openpyxl.load_workbook(trans_file, read_only=True)
    ws = wb.active
    data = []
    row_iter = ws.iter_rows(min_row=TRANS_START_ROW, min_col=TRANS_SOURCE_COL, max_col=TRANS_SOURCE_COL, values_only=True)
    for row in row_iter:
        data.append(row[0])
    wb.close()
    return data


def write_to_xl(trans_out):
    wb = openpyxl.load_workbook(trans_file)
    ws = wb.active
    for i in range(len(trans_out)):
        ws.cell(i + TRANS_START_ROW, TRANS_TARGET_COL).value = trans_out[i]
    wb.save(trans_file[:-5] + '_translated.xlsx')
    wb.close()

def getColIndex(letter):
    return ord(letter.lower()) - ord('a') + 1

if __name__ == '__main__':
    from tkinter import Tk
    from tkinter.filedialog import askopenfilename
    Tk().withdraw()

    # Prompt for input
    print('请选择要翻译的Excel文件:')
    trans_file = askopenfilename()

    target_lang = prompt_for_target_lang()

    TRANS_SOURCE_COL = getColIndex(input("请键入原文列（字母）:"))
    TRANS_TARGET_COL = getColIndex(input("请键入输出列："))
    TRANS_START_ROW = int(input("请键入需要翻译的第一格的行数（例:4）："))

    trans_in = load_xl()
    write_to_xl(translate_list(target_lang, trans_in))
