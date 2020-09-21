import openpyxl
import os
import random
import math
from statistics import mean
from ortools.linear_solver import pywraplp

def solve(cost, max_assignable):
    solver = pywraplp.Solver('SolveAssignmentProblem',
                           pywraplp.Solver.CBC_MIXED_INTEGER_PROGRAMMING)

    num_workers = len(cost)
    num_tasks = len(cost[0])
    x = {}

    for i in range(num_workers):
        for j in range(num_tasks):
            x[i, j] = solver.IntVar(0, 1, 'x[%i,%i]' % (i, j))

    for i in range(num_workers):
        solver.Add(solver.Sum([x[i, j] for j in range(num_tasks)]) <= max_assignable[i])
    for j in range(num_tasks):
        solver.Add(solver.Sum([x[i, j] for i in range(num_workers)]) == 1)

    solver.Minimize(solver.Sum([cost[i][j] * x[i,j] for i in range(num_workers)
                                                  for j in range(num_tasks)]))
    solver.Solve()

    net_cost = solver.Objective().Value()
    assignment = []

    for j in range(num_tasks):
        for i in range(num_workers):
            if x[i, j].solution_value() > 0:
                assignment.append(i)
                break

    return (net_cost, assignment)


def fit(values, sizes, distribution):
    solver = pywraplp.Solver('',
                           pywraplp.Solver.CBC_MIXED_INTEGER_PROGRAMMING)

    num_values = len(values)
    num_splits = len(sizes)

    x = {}

    for i in range(num_splits):
        for j in range(num_values):
            x[i, j] = solver.IntVar(0, 1, 'x[%i,%i]' % (i, j))

    infinity = solver.infinity()
    deviation = []

    for i in range(num_splits):
        deviation.append(solver.IntVar(0.0, infinity, 'dev[' + str(i) + ']'))

    for i in range(num_splits):
        solver.Add(solver.Sum([x[i, j] for j in range(num_values)]) == sizes[i])

    for j in range(num_values):
        solver.Add(solver.Sum([x[i, j] for i in range(num_splits)]) == 1)

    for i in range(num_splits):
        solver.Add(deviation[i] >= solver.Sum([x[i, j] * values[j]
                                    for j in range(num_values)]) - distribution[i])
        solver.Add(deviation[i] >= solver.Sum([x[i, j] * (-values[j])
                                    for j in range(num_values)]) + distribution[i])

    solver.Minimize(solver.Sum(deviation))
    solver.Solve()

    result = []
    for j in range(num_values):
        for i in range(num_splits):
            if x[i, j].solution_value() > 0:
                result.append(i)
                break

    return result


class Person:
    def __init__(self, Name, SENIORITY, MAX_ASSIGNABLE, REGION, GROUP, TBD1, TBD2, TBD3, TBD4):
        self.Name = Name
        self.SENIORITY = SENIORITY
        self.MAX_ASSIGNABLE = MAX_ASSIGNABLE
        self.REGION = REGION
        self.GROUP = GROUP
        self.TBD1 = TBD1
        self.TBD2 = TBD2
        self.TBD3 = TBD3
        self.TBD4 = TBD4

        self.difficulty = 0
        self.companies = set()


class Company:
    def __init__(self, ID, DIFFICULTY, GROUP, REGION, OLD_ASSIGNMENT, MANUAL_ASSIGNMENT, TBD1, TBD2, TBD3, TBD4):
        self.ID = ID
        self.DIFFICULTY = DIFFICULTY
        self.GROUP = GROUP
        self.REGION = REGION
        self.OLD_ASSIGNMENT = OLD_ASSIGNMENT
        self.MANUAL_ASSIGNMENT = MANUAL_ASSIGNMENT
        self.TBD1 = TBD1
        self.TBD2 = TBD2
        self.TBD3 = TBD3
        self.TBD4 = TBD4

        self.assignment = None
        self.shufflable = None

    def determine_shufflable(self, cluster_people):
        if self.MANUAL_ASSIGNMENT != None:
            self.shufflable = False
            return
        cost = cost_function(self.assignment, self)
        for person in cluster_people:
            if cost_function(person, self) - cost > 1e-6:
                if random.random() < FAIRNESS_SETTING:
                    self.shufflable = True
                    return
                self.shufflable = False
                return
        self.shufflable = True
        return


DIFFICULTY_COST = 0
REGION_COST = 0
GROUP_COST = 0
NEW_ASSIGNMENT_COST = 0
TBD1_COST = 0
TBD2_COST = 0
TBD3_COST = 0
TBD4_COST = 0

NA_DIFFICULTY_NUMBER = 0
FAIRNESS_SETTING = 0.0


def load_sheet(data_wb, sname, c_class, max_col):
    data_ws = data_wb[sname]
    data = []
    row_iter = data_ws.iter_rows(min_row=2, max_col=max_col, values_only=True)
    for row in row_iter:
        if row[0] != None:
            data.append(c_class(*row))
    return data


def load_consts(data_wb):
    global DIFFICULTY_COST, REGION_COST, GROUP_COST, NEW_ASSIGNMENT_COST, TBD1_COST, TBD2_COST, TBD3_COST, TBD4_COST
    global NA_DIFFICULTY_NUMBER, FAIRNESS_SETTING
    consts_ws = data_wb['程序参数']
    costs_iter = consts_ws.iter_rows(min_row=2, max_row=9, min_col=2, max_col=2, values_only=True)
    DIFFICULTY_COST = next(costs_iter)[0]
    REGION_COST = next(costs_iter)[0]
    GROUP_COST = next(costs_iter)[0]
    NEW_ASSIGNMENT_COST = next(costs_iter)[0]
    TBD1_COST = next(costs_iter)[0]
    TBD2_COST = next(costs_iter)[0]
    TBD3_COST = next(costs_iter)[0]
    TBD4_COST = next(costs_iter)[0]
    config_iter = consts_ws.iter_rows(min_row=2, max_row=3, min_col=4, max_col=4, values_only=True)
    NA_DIFFICULTY_NUMBER = next(config_iter)[0]
    FAIRNESS_SETTING = next(config_iter)[0]


def load_xl(data_fname):
    data_wb = openpyxl.load_workbook(data_fname, read_only=True)
    people = load_sheet(data_wb, '复核老师信息', Person, 9)
    companies = load_sheet(data_wb, '经销商信息', Company, 10)
    load_consts(data_wb)
    data_wb.close()
    return people, companies


def sanitize_rg_values(company, rg_attrib):
    rg_value = getattr(company, rg_attrib)
    if type(rg_value) == int or rg_value == '#N/A':
        setattr(company, rg_attrib, None)


def add_rg_cost(person_v, company_v):
    if company_v == None:
        return 1
    if person_v == company_v:
        return 0
    return 1


def cost_function(person, company):
    cost = 0
    cost += DIFFICULTY_COST * abs(company.DIFFICULTY - person.difficulty)
    cost += REGION_COST * add_rg_cost(person.REGION, company.REGION)
    cost += GROUP_COST * add_rg_cost(person.GROUP, company.GROUP)
    cost += NEW_ASSIGNMENT_COST * (person.Name != company.OLD_ASSIGNMENT)
    cost += TBD1_COST * add_rg_cost(person.TBD1, company.TBD1)
    cost += TBD2_COST * add_rg_cost(person.TBD2, company.TBD2)
    cost += TBD3_COST * add_rg_cost(person.TBD3, company.TBD3)
    cost += TBD4_COST * add_rg_cost(person.TBD4, company.TBD4)
    return cost


def min_max(xs):
    min, max = math.inf, -math.inf
    for x in xs:
        if x < min:
            min = x
        if x > max:
            max = x
    return min, max


if __name__ == '__main__':
    from tkinter import Tk
    from tkinter.filedialog import askopenfilename
    Tk().withdraw()

    data_fname = askopenfilename()
    people, companies = load_xl(data_fname)

    # Sanitize companies
    for company in companies:
        # Sanitize difficulty
        if type(company.DIFFICULTY) != int:
            company.DIFFICULTY = NA_DIFFICULTY_NUMBER
        # Sanitize RG values
        rg_attribs = ['GROUP', 'REGION', 'TBD1', 'TBD2', 'TBD3', 'TBD4']
        for rg_attrib in rg_attribs:
            sanitize_rg_values(company, rg_attrib)

    # index person objects by .Name
    person_by_name = {}
    for person in people:
        person_by_name[person.Name] = person

    # Clean up OLD_ASSIGNMENT
    # Account for manually assigned
    for company in companies:
        if company.OLD_ASSIGNMENT not in person_by_name:
            company.OLD_ASSIGNMENT = None
        if company.MANUAL_ASSIGNMENT != None:
            person_by_name[company.MANUAL_ASSIGNMENT].MAX_ASSIGNABLE -= 1

    # Clusterize people
    people_sorted = people[:]
    people_sorted.sort(key=lambda p : p.SENIORITY, reverse=True)
    cluster_peoples = []
    partition_sizes = []
    cur_sen = -1
    for person in people_sorted:
        if person.SENIORITY != cur_sen:
            cluster_peoples.append([])
            partition_sizes.append(0)
            cur_sen = person.SENIORITY
        cluster_peoples[-1].append(person)
        partition_sizes[-1] += person.MAX_ASSIGNABLE

    # Assign difficulties
    companies_nma_diffs = [
        company.DIFFICULTY
        for company in companies
        if company.MANUAL_ASSIGNMENT == None
    ]
    companies_nma_diffs.sort(reverse=True)

    st = 0
    for size, cluster_people in zip(partition_sizes, cluster_peoples):
        sp = min(st + size, len(companies_nma_diffs))
        average_difficulty = mean(companies_nma_diffs[st:sp])
        for person in cluster_people:
            person.difficulty = average_difficulty
        st = sp

    # Compute unfair assignments
    cost, max_assignable = [], []
    for person in people:
        max_assignable.append(person.MAX_ASSIGNABLE)
        t = []
        for company in companies:
            if company.MANUAL_ASSIGNMENT == None:
                t.append(cost_function(person, company))
        cost.append(t)
    net_cost, assignment = solve(cost, max_assignable)
    j = 0
    for company in companies:
        if company.MANUAL_ASSIGNMENT != None:
            company.assignment = person_by_name[company.MANUAL_ASSIGNMENT]
        else:
            company.assignment = people[assignment[j]]
            j += 1
        company.assignment.companies.add(company)

    # Clusterize companies based on unfair assignment
    cluster_companiess = []
    for cluster_people in cluster_peoples:
        cluster_companiess.append([])
        for person in cluster_people:
            for company in person.companies:
                cluster_companiess[-1].append(company)

    # Find shufflable companies with respect to their cluster
    for cluster_people, cluster_companies in zip(cluster_peoples, cluster_companiess):
        for company in cluster_companies:
            company.determine_shufflable(cluster_people)

    # Tally for parameters for fit(), and reassign for fairness
    for cluster_people, cluster_companies in zip(cluster_peoples, cluster_companiess):
        if len(cluster_people) == 1:
            continue

        cluster_ambiv_companies = [company for company in cluster_companies if company.shufflable]

        cluster_ambiv_companies_diff = [company.DIFFICULTY for company in cluster_ambiv_companies]

        cluster_people_ambiv_count = [
            sum(company.shufflable for company in person.companies)
            for person in cluster_people
        ]

        cluster_companies_diff_avg = mean(company.DIFFICULTY for company in cluster_companies)
        cluster_people_ideal_v = [
            int((cluster_companies_diff_avg * len(person.companies)
            - sum(company.DIFFICULTY for company in person.companies
            if not company.shufflable)))
            for person in cluster_people
        ]

        new_ambiv_assignment = fit(
            cluster_ambiv_companies_diff,
            cluster_people_ambiv_count,
            cluster_people_ideal_v
        )

        for idx, company in zip(new_ambiv_assignment, cluster_ambiv_companies):
            company.assignment.companies.remove(company)
            company.assignment = cluster_people[idx]
            company.assignment.companies.add(company)

    # Output
    output_wb = openpyxl.Workbook()

    assignments_ws = output_wb.active
    assignments_ws.title = 'Assignments'
    assignments_ws.cell(1, 1).value = '编号'
    assignments_ws.cell(1, 2).value = '本期负责人'
    assignments_ws.cell(1, 3).value = '是否被手动分配'
    assignments_ws.cell(1, 4).value = '问题总数'
    assignments_ws.cell(1, 5).value = '所属集团'
    assignments_ws.cell(1, 6).value = '大区'
    assignments_ws.cell(1, 7).value = '上一期负责人'
    for i, company in enumerate(companies):
        assignments_ws.cell(2 + i, 1).value = company.ID
        assignments_ws.cell(2 + i, 2).value = company.assignment.Name
        assignments_ws.cell(2 + i, 3).value = '√' if company.MANUAL_ASSIGNMENT != None else None
        assignments_ws.cell(2 + i, 4).value = company.DIFFICULTY
        assignments_ws.cell(2 + i, 5).value = company.GROUP
        assignments_ws.cell(2 + i, 6).value = company.REGION
        assignments_ws.cell(2 + i, 7).value = company.OLD_ASSIGNMENT

    statistics_ws = output_wb.create_sheet(title='Statistics')
    statistics_ws.cell(1, 1).value = '姓名'
    statistics_ws.cell(1, 2).value = 'Seniority'
    statistics_ws.cell(1, 3).value = '分配量'
    statistics_ws.cell(1, 4).value = '问题数平均值'
    statistics_ws.cell(1, 5).value = '最容易/最难'
    statistics_ws.cell(1, 6).value = '和上一期匹配'
    statistics_ws.cell(1, 7).value = '集团匹配'
    statistics_ws.cell(1, 8).value = '待定1匹配'
    statistics_ws.cell(1, 9).value = '待定2匹配'
    statistics_ws.cell(1, 10).value = '待定3匹配'
    statistics_ws.cell(1, 11).value = '待定4匹配'
    for i, person in enumerate(people):
        statistics_ws.cell(2 + i, 1).value = person.Name
        statistics_ws.cell(2 + i, 2).value = person.SENIORITY
        statistics_ws.cell(2 + i, 3).value = len(person.companies)
        statistics_ws.cell(2 + i, 4).value = mean(company.DIFFICULTY for company in person.companies)
        statistics_ws.cell(2 + i, 5).value = '%d/%d' % min_max(company.DIFFICULTY for company in person.companies)
        count_in = lambda collection : sum(person.Name == company.OLD_ASSIGNMENT for company in collection)
        statistics_ws.cell(2 + i, 6).value = '%d/%d' % (count_in(person.companies), count_in(companies))
        rg_attribs = ['GROUP', 'TBD1', 'TBD2', 'TBD3', 'TBD4']
        for j, rg_attrib in enumerate(rg_attribs):
            count_in = lambda collection : sum(
                getattr(person, rg_attrib) == getattr(company, rg_attrib)
                for company in collection
            )
            if getattr(person, rg_attrib) != None:
                statistics_ws.cell(2 + i, 7 + j).value = '%d/%d' % (count_in(person.companies), count_in(companies))

    output_wb.save(os.path.join(os.path.dirname(data_fname), 'output.xlsx'))
    output_wb.close()
