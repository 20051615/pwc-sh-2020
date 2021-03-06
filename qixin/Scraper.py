import config
import openpyxl
from selenium import webdriver
from DriverOps import login, set_month_to_fetch, get_company_code, navigate_to_subpage, fetch_section, fetch_news_section, fetch_tax_section
from tkinter import Tk
from tkinter.filedialog import askopenfilename

SECTIONS = [
    "变更记录", 
    "裁判文书", 
    "被执行人", 
    "失信被执行", 
    "限制高消费", 
    "开庭公告", 
    "股权冻结", 
    "行政处罚", 
    "环保处罚", 
    "经营异常", 
    "税务异常", 
    "严重违法失信", 
    "动产抵押", 
    "股权出质", 
    "新闻舆情"
]

SPECIAL_SECTIONS = {"新闻舆情", "税务异常"}

SUBPAGE = {
    "变更记录": "企业概况",
    "裁判文书": "司法涉诉",
    "被执行人": "司法涉诉",
    "失信被执行": "司法涉诉",
    "限制高消费": "司法涉诉",
    "开庭公告": "司法涉诉",
    "股权冻结": "司法涉诉",
    "行政处罚": "经营预警",
    "环保处罚": "经营预警",
    "经营异常": "经营预警",
    "税务异常": "经营预警",
    "严重违法失信": "经营预警",
    "动产抵押": "经营信息",
    "股权出质": "经营信息",
    "新闻舆情": "经营信息"
}

def getColIndex(letter):
    return ord(letter.lower()) - ord('a') + 1

def _pair_extend(risk_contents, risk_types, new_risk_contents, section_name):
    risk_contents[0].extend(new_risk_contents[0])
    risk_contents[1].extend(new_risk_contents[1])
    risk_types[0].extend([section_name] * len(new_risk_contents[0]))
    risk_types[1].extend([section_name] * len(new_risk_contents[1]))

def getRiskStrings(company_name):
    company_code = get_company_code(driver, company_name)
    if company_code is None:
        return "未找到", "！！！未找到全名完全匹配的经销商！！！"
    risk_contents = [], []
    risk_types = [], []
    current_subpage = ""
    current_subpage_sections = None
    for section_name in SECTIONS:
        extend_result_with = lambda new_risk_contents : _pair_extend(risk_contents, risk_types, new_risk_contents, section_name)

        if SUBPAGE[section_name] != current_subpage:
            current_subpage = SUBPAGE[section_name]
            current_subpage_sections = navigate_to_subpage(driver, company_code, current_subpage)
        section = current_subpage_sections[section_name]

        if section_name not in SPECIAL_SECTIONS:
            main_table = fetch_section(driver, section, section_name)
            if main_table is not None:
                extend_result_with(main_table.to_entrys())
        else:
            if section_name == "税务异常":
                tab1_table, tab2_table = fetch_tax_section(driver, section)
                if tab1_table is not None:
                    extend_result_with(tab1_table.to_entrys())
                if tab2_table is not None:
                    extend_result_with(tab2_table.to_entrys())
            elif section_name == "新闻舆情":
                news_titles = fetch_news_section(driver, section)
                if news_titles is not None:
                    extend_result_with(news_titles)

    format_to_excel_string = lambda strings_pair : "\n".join(
        "【" + str(idx + 1) + "】" + string
        for idx, string in enumerate(strings_pair[0])
    ) + "\n" if strings_pair[0] and strings_pair[1] else "" + "\n".join(
        "【日期未经核实筛选_" + str(idx + 1) + "】" + string
        for idx, string in enumerate(strings_pair[1])
    )
    return format_to_excel_string(risk_types), format_to_excel_string(risk_contents)

if __name__ == "__main__":
    Tk().withdraw()
    excel_file = askopenfilename(title="Select excel sheet to process")

    SECTION_NAME_COL = getColIndex(input("请键入要填入的风险类别列（例:K）："))
    CONTENT_COL = getColIndex(input("请键入要填入的风险内容列："))
    COMPANY_COL = getColIndex(input("请键入要读取的经销商名称列："))
    START_ROW = int(input("请键入从哪一行开始爬取（例:3）："))
    
    if input("是否手动选择获取哪个月的信息？(y/n)：").lower() == 'y':
        print("请留意，获取过旧的信息可能会导致程序遗漏本应获取到的内容。")
        set_month_to_fetch(
                           int(input("年（例:2020）：")),
                           int(input("月（例:6）："))
                           )
    
    wb = openpyxl.load_workbook(excel_file, read_only=True)
    company_names = [
        row[0]
        for row in wb.active.iter_rows(min_row=START_ROW, min_col=COMPANY_COL, max_col=COMPANY_COL, values_only=True)
    ]
    wb.close()

    with webdriver.Chrome() as driver:
        login(driver, config.USERNAME, config.PASSWORD)
        
        wb = openpyxl.load_workbook(excel_file)
        for idx, company_name in enumerate(company_names):
            section_name_string, content_string = getRiskStrings(company_name)
            
            wb.active.cell(START_ROW + idx, SECTION_NAME_COL).value = section_name_string
            wb.active.cell(START_ROW + idx, CONTENT_COL).value = content_string
            wb.save(excel_file[:-5] + "_filled.xlsx")
        
        wb.close()